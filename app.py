import tkinter as tk
from datetime import date
import openpyxl as xl
import random


def on_yes_click():
    message = "Sorry, no reward for you today. Try again tomorrow."
    update_rewards_excel(0)
    calculate_total_rewards(message)


def on_no_click():
    workbook = xl.load_workbook("money_rewards.xlsx")
    sheet = workbook.active
    max_row = sheet.max_row
    random_row = random.randint(1, max_row)
    value = sheet[f'A{random_row}'].value
    message = f"Today, you earned ${value} toward your goal. Congrats!"
    update_rewards_excel(value)
    calculate_total_rewards(message)


def update_rewards_excel(value):
    workbook = xl.load_workbook("rewards.xlsx")
    sheet = workbook.active
    sheet.append([date.today(), value])
    workbook.save("rewards.xlsx")


def calculate_total_rewards(message):
    workbook = xl.load_workbook("rewards.xlsx")
    sheet = workbook.active
    values = [int(cell.value)
              for cell in sheet['B'] if isinstance(cell.value, int)]
    total_rewards = sum(values)
    message_label.config(text=f"{message}\nTotal rewards: ${total_rewards}")


window = tk.Tk()
window.title("Scott's Treats App")
window.geometry("600x300")

title_label = tk.Label(window, text="Did you eat any treats today?")
title_label.pack()

date_label = tk.Label(window, text=date.today())
date_label.pack()

yes_button = tk.Button(window, text="Yes", command=on_yes_click)
yes_button.pack()

no_button = tk.Button(window, text="No", command=on_no_click)
no_button.pack()

message_label = tk.Label(window, text="")
message_label.pack()

window.mainloop()
