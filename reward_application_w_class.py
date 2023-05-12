import openpyxl as op
import random as r
import datetime as dt
import tkinter as tk
import tkinter.font as font
from tkinter import ttk
from windows import set_dpi_awareness



set_dpi_awareness()

class TreatTracker(tk.Tk):
    def __init__(self):
        super().__init__()
    
        self.title("Treat Tracker")
        
        self.columnconfigure(0, weight=1)
        self.rowconfigure((1,2), weight=1)
        
        ttk.Label(self, text="Did you eat any treats today?", padding=(30, 10)).grid()

        yesNoFrame(self).grid()
        quitButtonFrame(self).grid()


class yesNoFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)

        self.grid(sticky="nsew")

        self.columnconfigure((0,1), weight=1)
        self.rowconfigure(0, weight=1)

        yes_button = ttk.Button(self, command=self.yes, text="Yes")
        no_button = ttk.Button(self, command=self.no, text="No")
        
        yes_button.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        no_button.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        

    def yes(self):
        # Find the next empty cell
        next_reward_cell = nextEmptyCell(2, 1)    
        # Place the reward of 0 into the cell
        next_reward_cell.value = 0
        sum_of_row = sumRow()
        print(sum_of_row)
        # Save the relevant excel file
        wb.save("Rewards1.xlsx")
    
    def no(self):
        #Find the empty cell
        next_reward_cell = nextEmptyCell(2, 1)
        # Open the workbooks with the money rewards
        rewb = op.load_workbook("money_rewards.xlsx")
        # Activate the first sheet
        ws = rewb.active   
        # Choose a reward value and erase it from the rewards excel document
        while True:
            row_number = r.randint(1, 41)
            cell = f"A{row_number}"
            reward_amount = ws[f"{cell}"].value
            if reward_amount is not None:
                ws[f"{cell}"].value = None
                break
            else:
                continue
        # Place the reward amount value in the previously empty cell
        next_reward_cell.value = reward_amount

        sum_of_row = sumRow()
        print(sum_of_row)

        # Update relevant excel documents
        wb.save("Rewards1.xlsx")
        rewb.save("money_rewards.xlsx")

class quitButtonFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)

        self.grid(sticky="nsew")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        quit_button = ttk.Button(self, command=container.destroy, text="Quit")
        quit_button.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)


# A function to find the next empty cell
def nextEmptyCell(row, column):
    next_cell = ws.cell(row=row, column=column)
    while next_cell.value is not None:
        next_cell = ws.cell(row=row, column=next_cell.column+1)
    
    return next_cell

# A function to sum the values of a row
def sumRow():
    row_values = [cell.value if cell.value is not None else 0 for cell in ws[2]]
    row_sum = sum(row_values)

    return row_sum

# Load or create the workbook
try:
    wb = op.load_workbook("Rewards1.xlsx")
except FileNotFoundError:
    wb = op.Workbook()
    wb.save("Rewards1.xlsx")

# Select the first worksheet
ws = wb.active

# Find the next available cell in the first row
next_date_cell = nextEmptyCell(1, 1)

# Write the current date in the cell
next_date_cell.value = dt.date.today()


root = TreatTracker()
font.nametofont("TkDefaultFont").configure(size=15)
root.mainloop()

