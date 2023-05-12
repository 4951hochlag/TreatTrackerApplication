import openpyxl as op
import random as r
import datetime as dt
import tkinter as tk
import tkinter.font as font
from tkinter import ttk
from windows import set_dpi_awareness

set_dpi_awareness()


# A function to find the next empty cell
def nextEmptyCell(row, column):
    next_cell = ws.cell(row=row, column=column)
    while next_cell.value is not None:
        next_cell = ws.cell(row=row, column=next_cell.column+1)
    
    return next_cell

# A function to sum the values of a row
def sumRow():
    row_values = [cell.value if cell.value is not None else 0 for cell in ws[2]]
    row_sum = sum(row_values)

    return row_sum

# The command function for the yes button
def yes():
    next_reward_cell = nextEmptyCell(2, 1)    
    next_reward_cell.value = 0
    wb.save("Rewards1.xlsx")

# The command function for the no button
def no():
    
    # Find the next empty cell in the row
    next_reward_cell = nextEmptyCell(2, 1)
    
    # Open the workbooks with the money rewards
    rewb = op.load_workbook("money_rewards.xlsx")

    # Activate the first sheet
    ws = rewb.active   
    
    # Choose a reward value and erase it
    while True:
        row_number = r.randint(1, 41)
        cell = f"A{row_number}"
        reward_amount = ws[f"{cell}"].value
        if reward_amount is not None:
            ws[f"{cell}"].value = None
            break
        else:
            continue
    
    # Place the reward amount value in the previously empty cell
    next_reward_cell.value = reward_amount


    # Update relevant excel documents
    wb.save("Rewards1.xlsx")
    rewb.save("money_rewards.xlsx")

# A function to load a workbook



# Load or create the workbook
try:
    wb = op.load_workbook("Rewards1.xlsx")
except FileNotFoundError:
    wb = op.Workbook()
    wb.save("Rewards1.xlsx")

# Select the first worksheet
ws = wb.active

# Find the next available cell in the first row
next_date_cell = nextEmptyCell(1, 1)

# Write the current date in the cell
next_date_cell.value = dt.date.today()


root = tk.Tk()
root.title("Treat Tracker")

font.nametofont("TkDefaultFont").configure(size=15)

root.columnconfigure(0, weight=1)
root.rowconfigure((1,2), weight=1)


ttk.Label(root, text="Did you eat any treats today?", padding=(30, 10)).grid()

yes_no_buttons = ttk.Frame(root, padding=(15, 0))
yes_no_buttons.grid(sticky="nsew")

yes_no_buttons.columnconfigure((0, 1), weight=1)
yes_no_buttons.rowconfigure(0, weight=1)

yes_button = ttk.Button(yes_no_buttons, text="Yes", command=yes)
yes_button.grid(row=0, column=0, sticky="nsew") 

no_button = ttk.Button(yes_no_buttons, text="No", command=no)
no_button.grid(row=0, column=1, sticky="nsew")

quit_button_frame = ttk.Frame(root, padding=(15, 5))
quit_button_frame.grid(sticky="nsew")

quit_button_frame.columnconfigure(0, weight=1)
quit_button_frame.rowconfigure(0, weight=1)

quit_button = ttk.Button(quit_button_frame, text="Quit", command=root.destroy)
quit_button.grid(row=0, column=0, sticky="nsew")

root.mainloop()

