import openpyxl as op
import random as r
import datetime as dt
import tkinter as tk
import tkinter.font as font
from tkinter import ttk
from rewapplication import nextEmptyCell, sumRow
from windows import set_dpi_awareness

set_dpi_awareness()


# Class for the application window
class TreatTracker(tk.Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
    
        self.title("Treat Tracker")
              
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        
        OpeningFrame(self).grid()
        


class OpeningFrame(ttk.Frame):
    def __init__(self, container, *kwargs):
        super().__init__(container, *kwargs)

        # Identify the active worksheet
        self.ws1 = wb.active    

        self.grid(sticky="nsew")

        # Row and column configurations
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        

        # Widgets
        question = ttk.Label(self, text="Did you eat any treats today?")
        yes_button = ttk.Button(self, command=self.yesButton, text="Yes")
        no_button = ttk.Button(self, command=self.noButton, text="No")
        quit_button = ttk.Button(self, command=container.destroy, text="Quit")
        

        # Layout
        question.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        yes_button.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        no_button.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        quit_button.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)

        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)
        

    def yesButton(self):  
        # Find the next empty cell
        next_reward_cell = nextEmptyCell(self.ws1, 2, 1)    
        # Place the reward of 0 into the cell
        next_reward_cell.value = 0
        # Save the relevant excel file
        wb.save("Rewards1.xlsx")
        
        # Sum the rewards row
        sum_row = sumRow(self.ws1)
        print(sum_row)
        
        
    
    def noButton(self):
        #Find the empty cell
        next_reward_cell = nextEmptyCell(self.ws1, 2, 1)
        # Open the workbooks with the money rewards
        rewb = op.load_workbook("money_rewards.xlsx")
        # Activate the first sheet
        ws2 = rewb.active   
        # Choose a reward value and erase it from the rewards excel document
        while True:
            row_number = r.randint(1, 41)
            cell = f"A{row_number}"
            reward_amount = ws2[f"{cell}"].value
            if reward_amount is not None:
                ws2[f"{cell}"].value = None
                break
            else:
                continue
        # Place the reward amount value in the previously empty cell
        next_reward_cell.value = reward_amount
        
        # Update relevant excel documents
        wb.save("Rewards1.xlsx")
        rewb.save("money_rewards.xlsx")
        
        # Sum the rewards row
        sum_row = sumRow(self.ws1)
        print(sum_row)
        

# Load or create the workbook
try:
    wb = op.load_workbook("Rewards1.xlsx")
except FileNotFoundError:
    wb = op.Workbook()
    wb.save("Rewards1.xlsx")

# Select the first worksheet
ws1 = wb.active

# Find the next available cell in the first row
next_date_cell = nextEmptyCell(ws1, 1, 1)

# Write the current date in the cell
next_date_cell.value = dt.date.today()


root = TreatTracker()
font.nametofont("TkDefaultFont").configure(size=15)
root.mainloop()

