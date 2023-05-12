import random as r
import os as o
import datetime as dt
import openpyxl as op
import datetime as dt
import tkinter as tk
import tkinter.font as font
from tkinter import ttk
from rewapplication import nextEmptyCell
from windows import set_dpi_awareness

set_dpi_awareness()


# Class for the application window
class TreatTracker(tk.Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
    
        self.title("Treat Tracker")
        self.frames = dict()

        container = ttk.Frame(self)
        container.grid(padx=40, pady=40, sticky="nsew")
        
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        
        no_frame = NoFrame(container, self)
        no_frame.grid(row=0, column=0, sticky="nsew")

        yes_frame = YesFrame(container, self)
        yes_frame.grid(row=0, column=0, sticky="nsew")

        opening_frame = OpeningFrame(container, self)
        opening_frame.grid(row=0, column=0, sticky="nsew")

        self.frames[OpeningFrame] = opening_frame 
        self.frames[NoFrame] = no_frame
        self.frames[YesFrame] = yes_frame
          
        self.show_frame(OpeningFrame)
       
    def show_frame(self, container):
        frame = self.frames[container]
        frame.tkraise()


class OpeningFrame(ttk.Frame):
    def __init__(self, container, controller, *kwargs):
        super().__init__(container, *kwargs)

        # Identify the active worksheet
        self.ws1 = wb.active  

        self.grid(sticky="nsew")

        # Row and column configurations
        self.columnconfigure((0,1), weight=1)
        self.rowconfigure((0,1,2), weight=1)

        # Widgets
        question = ttk.Label(self, text="Did you eat any treats today?")
        yes_button = ttk.Button(self, command=lambda: controller.show_frame(YesFrame), text="Yes")
        no_button = ttk.Button(self, command= lambda: controller.show_frame(NoFrame), text="No")
        quit_button = ttk.Button(self, command=container.destroy, text="Quit")
        

        # Layout
        question.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        yes_button.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        no_button.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        quit_button.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
 
 
class NoFrame(ttk.Frame):
    def __init__(self, container, controller, *kwargs):
        super().__init__(container, *kwargs)

        self.grid(sticky="nsew")
        
        # Initialize the reward variable
        self.reward_amount = 0

        # Identify the active worksheet
        self.ws1 = wb.active  
        
        self.determineReward()
        self.storeRewardsum(self.sumRow())
        

        congrats = ttk.Label(self,text=f"Today, you earned ${self.reward_amount} " 
                             "toward your goal!\n Congratulations!!!")
        new_total = ttk.Label(self, text=f"You have earned a total of ${self.sumRow()} toward your goal")
        graph_button = ttk.Button(self, command=self.showGraph, text="See your progress!")
        excel_button = ttk.Button(self, command=self.showExcel, text="See spreadsheet")
        quit_button = ttk.Button(self, command=container.destroy, text="Quit")

        congrats.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        new_total.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        graph_button.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
        excel_button.grid(row=2, column=1, sticky="nsew", padx=5, pady=5)
        quit_button.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)

    def determineReward(self):
        #Find the empty cell
        next_reward_cell = nextEmptyCell(self.ws1, 2, 1)
        # Open the workbooks with the money rewards
        rewb = op.load_workbook("money_rewards.xlsx")
        # Activate the first sheet
        ws2 = rewb.active   
        # Choose a reward value and erase it from the rewards excel document
        while True:
            row_number = r.randint(1, 41)
            cell = f"A{row_number}"
            self.reward_amount = ws2[f"{cell}"].value
            if self.reward_amount is not None:
                ws2[f"{cell}"].value = None
                break
            else:
                continue
        # Place the reward amount value in the previously empty cell
        next_reward_cell.value = self.reward_amount
        
        # Update relevant excel documents
        wb.save("Rewards1.xlsx")
        rewb.save("money_rewards.xlsx")
        
        return self.reward_amount

    def sumRow(self):

        row_values = [cell.value if cell.value is not None else 0 for cell in self.ws1[2]]
        row_sum = sum(row_values)
        return row_sum
    
    def showGraph(self):
        pass

    def showExcel(self):
        o.system(r"Rewards1.xlsx")

    def storeRewardsum(self, new_sum):
        wb = op.load_workbook("PastSums.xlsx")
        ws3 = wb.active
        new_sum_cell = nextEmptyCell(ws3, 2, 1)
        new_sum_cell.value = new_sum
        wb.save("PastSums.xlsx")

class YesFrame(ttk.Frame):
    def __init__(self, container, controller, *kwargs):
        super().__init__(container, *kwargs)

        # Identify the active worksheet
        self.ws1 = wb.active    

        self.grid(sticky="nsew")

        self.zeroRewards()

        sorry = ttk.Label(self,text=f"Sorry! No reward for you today! Try again tomorrow.") 
        new_total = ttk.Label(self, text=f"You have earned a total of ${self.sumRow()} toward your goal")
        graph_button = ttk.Button(self, command=self.showGraph, text="See your progress!")
        excel_button = ttk.Button(self, command=self.showExcel, text="See spreadsheet")
        quit_button = ttk.Button(self, command=container.destroy, text="Quit")

        sorry.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        new_total.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        graph_button.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
        excel_button.grid(row=2, column=1, sticky="nsew", padx=5, pady=5)
        quit_button.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)

    def zeroRewards(self):
        # Find the next empty cell
        next_reward_cell = nextEmptyCell(self.ws1, 2, 1)    
        # Place the reward of 0 into the cell
        next_reward_cell.value = 0
        # Save the relevant excel file
        wb.save("Rewards1.xlsx")

    def sumRow(self):

        row_values = [cell.value if cell.value is not None else 0 for cell in self.ws1[2]]
        row_sum = sum(row_values)
      
        return row_sum
    
    def showGraph(self):
        pass

    def showExcel(self):
        o.system(r"Rewards1.xlsx")


# Load or create the workbooks
wb = op.load_workbook("Rewards1.xlsx")
sumwb = op.load_workbook("PastSums.xlsx")

ws1 = wb.active
ws4 = sumwb.active

# Find the next available cell in the first row
next_date_cell_1 = nextEmptyCell(ws1, 1, 1)
next_date_cell_2 = nextEmptyCell(ws4, 1, 1)

# Write the current date in the cell
next_date_cell_1.value = dt.date.today()
next_date_cell_2.value = dt.date.today()


root = TreatTracker()
font.nametofont("TkDefaultFont").configure(size=15)
root.mainloop()

